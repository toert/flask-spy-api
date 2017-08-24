import requests
import urllib.parse
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
import csv
import io
import http.client
from datetime import datetime
from pytz import timezone

http.client._MAXHEADERS = 1000

LAST_UPDATE_TIME = '- никогда'
LOGIN = 'ivan-porch@bk.ru'
PASSWORD = '5379200f47735ff99df8eae75ac98845'
API_URL = 'http://api.spywords.ru'
LIMIT = '3'


def open_file_as_lines(filepath):
    with open(filepath, 'r') as data:
        return [str(line) for line in data]


def get_unique_links(sites):
    return list(set([site['url'] for site in sites]))


def make_request(url_without_schema):
    site_url = 'http://{}'.format(url_without_schema)
    try:
        raw_html = requests.get(site_url, timeout=5.0).text
    except:
        site_url = 'https://{}'.format(url_without_schema)
        try:
            raw_html = requests.get(site_url, timeout=5.0).text
        except:
            return None
    return raw_html


def check_presence_text_in_site(raw_html, text):
    raw_html = raw_html.lower()
    return re.findall(text, raw_html) != []


def format_phone(entry_phone, result_phone_lenght=11):
    phone_without_symbols = ''.join(re.findall(r'\d', entry_phone))
    return phone_without_symbols[-result_phone_lenght:]


def collect_contacts(site, raw_html):
    site['e-mail'] = []
    site['phone'] = []
    soup = BeautifulSoup(raw_html, 'html.parser')
    for a in soup.find_all('a', href=True):
        if re.findall(r'mailto:.', a['href']):
            site['e-mail'].append(re.findall(r'mailto:(.+)', a['href'])[0])
        if re.findall(r'tel:.', a['href']):
            site['phone'].append(format_phone(re.findall(r'tel:(.+)', a['href'])[0]))
    if not site['phone']:
        site['phone'].extend(re.findall(r'(\((?:495|499)\) \d{3}\-\d{2}\-\d{2})', raw_html))
    site['phone'] = list(set(site['phone']))
    site['e-mail'] = list(set(site['e-mail']))
    return site


def encode_params(params):
    encoded_params = {}
    for key, value in params.items():
        encoded_params[urllib.parse.quote(key, encoding='cp1251')] = urllib.parse.quote(value, encoding='cp1251')
    return encoded_params


def get_url(url, params):
    url = url + '?'
    for key, value in params.items():
        url = '{}{}={}&'.format(url, key, value)
    return url


def send_request_to_API(url, params):
    params = urllib.parse.urlencode(params, encoding='cp1251')
    url = '{}?{}'.format(url, params)
    return requests.get(url).text


def convert_respond_to_dict(text):
    rows = []
    indexes = [index for index in text.splitlines()[0].split('\t')]
    for line in text.splitlines()[1:]:
        row = {}
        for num_in_row, value in enumerate(line.split('\t')):
            row[indexes[num_in_row]] = value
        rows.append(row)
    return rows


def get_competitors(url, limit):
    raw_respond = send_request_to_API(API_URL, {
        'method': 'DomainAdvCompetitors',
        'site': url,
        'se': 'yandex',
        'limit': str(limit),
        'login': LOGIN,
        'token': PASSWORD,
    })
    return convert_respond_to_dict(raw_respond)


def parse_additional_info(site):
    raw_html = make_request(url_without_schema=site['Domain'])
    if raw_html is None:
        site['rtrg'] = ''
        site['connect.facebook'] = ''
        site['phone'] = ''
        site['e-mail'] = ''
        return site
    site['rtrg'] = check_presence_text_in_site(raw_html, 'rtrg')
    site['connect.facebook'] = check_presence_text_in_site(raw_html, 'connect.facebook')
    site = collect_contacts(site, raw_html)
    return site


def output_sites_info_to_xlsx(sites_info):
    wb = Workbook()
    sheet = wb.active
    sheet.cell(row=2, column=1).value = 'ID'
    sheet.cell(row=2, column=2).value = 'Тематика'
    sheet.cell(row=2, column=3).value = 'URL'
    sheet.cell(row=2, column=4).value = 'RTRG'
    sheet.cell(row=2, column=5).value = 'connect.facebook.net'
    sheet.cell(row=2, column=6).value = 'Номер телефона'
    sheet.cell(row=2, column=7).value = 'email'
    sheet.cell(row=2, column=8).value = 'Competition Level, %'
    sheet.cell(row=2, column=9).value = 'Кол-во пересекающихся ключевых слов'
    sheet.cell(row=2, column=10).value = 'Unique Keys'
    sheet.cell(row=2, column=11).value = 'Общее кол-во запросов домена в контекстной рекламе / поиске'
    sheet.cell(row=2, column=12).value = 'TotUniqAds'
    sheet.cell(row=2, column=13).value = 'Средняя позиция объявлений рекламодателя'
    sheet.cell(row=2, column=14).value = 'Оценка месячного трафика из контекстной рекламы'
    id = 0
    for keyword, list_of_sites in sites_info.items():
        list_of_sites = sorted(list_of_sites,
                               key=lambda x: float(x['AdTraf'].replace(' ', '')), reverse=True)
        for site in list_of_sites:
            sheet.cell(row=id + 3, column=1).value = id + 1
            sheet.cell(row=id + 3, column=2).value = keyword
            sheet.cell(row=id + 3, column=3).value = site['Domain']
            sheet.cell(row=id + 3, column=4).value = site['rtrg']
            sheet.cell(row=id + 3, column=5).value = site['connect.facebook']
            sheet.cell(row=id + 3, column=6).value = list_to_string(site['phone'])
            sheet.cell(row=id + 3, column=7).value = list_to_string(site['e-mail'])
            sheet.cell(row=id + 3, column=8).value = site['Competition Level, %']
            sheet.cell(row=id + 3, column=9).value = site['KeyOverlap']
            sheet.cell(row=id + 3, column=10).value = site['Unique Keys']
            sheet.cell(row=id + 3, column=11).value = site['KeysTot']
            sheet.cell(row=id + 3, column=12).value = site['TotUniqAds']
            sheet.cell(row=id + 3, column=13).value = site['AvgPos']
            sheet.cell(row=id + 3, column=14).value = site['AdTraf']
            id = id + 1
    wb.save('result.xlsx')
    return 'result.xlsx'


def output_sites_to_csv(sites_info):
    output_io = io.StringIO()
    fieldnames = ['id', 'keyword', 'Domain', 'rtrg', 'connect.facebook',
                  'phone', 'e-mail', 'Competition Level, %', 'KeyOverlap',
                  'Unique Keys', 'KeysTot', 'TotUniqAds', 'AvgPos', 'AdTraf']
    writer = csv.DictWriter(output_io, fieldnames=fieldnames)
    writer.writeheader()
    id = 1
    for keyword, list_of_sites in sites_info.items():
        list_of_sites = sorted(list_of_sites,
                               key=lambda x: float(x['AdTraf'].replace(' ', '')), reverse=True)
        for site in list_of_sites:
            site['id'] = id
            site['keyword'] = keyword
            site['phone'] = list_to_string(site['phone'])
            site['e-mail'] = list_to_string(site['e-mail'])
            writer.writerow(site)
            id = id + 1
    return output_io.getvalue().strip('\r\n')


def list_to_string(list):
    if not list:
        return ''
    result = list[0]
    for obj in list[1:]:
        if obj is not None:
            result = '{}, {}'.format(result, obj)
    return result


def parse_info(search_words, login, password, limit):
    keyword_info = {}
    for word in search_words:
        raw_respond = send_request_to_API(API_URL, {
            'method': 'KeywordAdv',
            'word': word,
            'se': 'yandex',
            'login': login,
            'token': password,
        })
        respond = convert_respond_to_dict(raw_respond)
        if respond:
            most_popular_site = max(respond, key=lambda site: int(site['KeysTot'].replace(' ', '')))
        else:
            continue
        keyword_info[word] = get_competitors(most_popular_site['Domain'], limit)
    for keyword, list_of_sites in keyword_info.items():
        for site in list_of_sites:
            site = parse_additional_info(site)
    return output_sites_info_to_xlsx(keyword_info)


if __name__ == "__main__":
    search_words = open_file_as_lines('words.txt')
    keyword_info = {}
    for word in search_words:
        raw_respond = send_request_to_API(API_URL, {
            'method': 'KeywordAdv',
            'word': word,
            'se': 'yandex',
            'login': LOGIN,
            'token': PASSWORD,
        })
        respond = convert_respond_to_dict(raw_respond)
        if respond:
            most_popular_site = max(respond, key=lambda site: int(site['KeysTot'].replace(' ', '')))
            print('API прислал {} сайтов по запросу {}, используем {}'
                  .format(len(respond), word, most_popular_site['Domain']))
        else:
            print('API ничего не прислал по запросу {}'.format(word))
            print('Код: {}'.format(raw_respond))
            continue
        keyword_info[word] = get_competitors(most_popular_site['Domain'])
    for keyword, list_of_sites in keyword_info.items():
        for site in list_of_sites:
            site = parse_additional_info(site)
    print(keyword_info)
    output_sites_info_to_xlsx(keyword_info)
